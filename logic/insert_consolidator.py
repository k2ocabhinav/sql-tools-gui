"""
INSERT Consolidator Logic Module
================================
Consolidates multiple INSERT statements from MySQL Workbench exports into 
a single INSERT statement per table, with proper table name extraction.

Features:
- Process SQL files from folder
- Modify files in-place or to separate output folder
- Generate combined file with TRUNCATE + INSERTs
- Track record counts for each table
- Generate Excel summary

Author: Abhinav Prasad
"""

import os
import re
from datetime import datetime
from pathlib import Path


def extract_table_name(content: str) -> str | None:
    """
    Extract table name from SQL comment.
    Handles:
    - SELECT * FROM table_name
    - SELECT col1, col2 FROM table_name
    - SELECT * FROM schema_name.table_name
    """
    # Pattern to match: -- Query: SELECT ... FROM [schema.]table_name
    pattern = r'--\s*Query:\s*SELECT\s+.+?\s+FROM\s+(?:\w+\.)?(\w+)'
    match = re.search(pattern, content, re.IGNORECASE)
    if match:
        return match.group(1)
    return None


def count_insert_statements(content: str) -> int:
    """Count the number of INSERT statements in content."""
    return len(re.findall(r'INSERT\s+INTO', content, re.IGNORECASE))


def extract_values_clause(insert_statement: str) -> str | None:
    """
    Extract the VALUES clause content (the tuple only) from an INSERT statement.
    Pattern: VALUES (...);
    Returns just the (...) part
    """
    match = re.search(r'VALUES\s*(\(.*\))\s*;\s*$', insert_statement, re.DOTALL)
    if match:
        return match.group(1)
    return None


def extract_column_definition(insert_statement: str) -> str | None:
    """
    Extract the column definition part from an INSERT statement.
    Pattern: INSERT INTO `` (`col1`,`col2`,...)
    Returns: (`col1`,`col2`,...)
    """
    # Match INSERT INTO `` (columns) - empty backticks
    match = re.search(r'INSERT\s+INTO\s+``\s*(\([^)]+\))\s*VALUES', insert_statement)
    if match:
        return match.group(1)
    return None


def process_single_content(raw_content: str) -> tuple[str, str | None, int]:
    """
    Process raw SQL content and return processed content.
    
    Returns: (processed_content, table_name, record_count)
    """
    # Extract table name from comment
    table_name = extract_table_name(raw_content)
    if not table_name:
        return raw_content, None, 0
    
    # Split content into lines and separate comment block from INSERT statements
    lines = raw_content.split('\n')
    
    comment_lines = []
    insert_lines = []
    comment_ended = False
    
    for line in lines:
        if not comment_ended:
            comment_lines.append(line)
            if '*/' in line:
                comment_ended = True
        else:
            insert_lines.append(line)
    
    # Find all INSERT statements with empty backticks
    insert_content = '\n'.join(insert_lines)
    
    # Pattern matches: INSERT INTO `` (`cols`) VALUES (...);
    insert_pattern = r'INSERT\s+INTO\s+``\s*\([^)]+\)\s*VALUES\s*\(.*?\);'
    insert_matches = re.findall(insert_pattern, insert_content, re.DOTALL)
    
    if not insert_matches:
        # No matches with empty backticks - return as is
        return raw_content, table_name, count_insert_statements(raw_content)
    
    # Extract column definition from first INSERT
    column_def = extract_column_definition(insert_matches[0])
    if not column_def:
        return raw_content, table_name, len(insert_matches)
    
    # Extract VALUES clauses from all INSERT statements
    values_list = []
    for insert in insert_matches:
        values = extract_values_clause(insert)
        if values:
            values_list.append(values)
    
    if not values_list:
        return raw_content, table_name, len(insert_matches)
    
    record_count = len(values_list)
    
    # Build the consolidated INSERT statement
    comment_block = '\n'.join(comment_lines)
    
    if len(values_list) == 1:
        output = f"{comment_block}\nINSERT INTO `{table_name}` {column_def} VALUES {values_list[0]};\n"
    else:
        # First value on same line, remaining with leading comma
        first_value = values_list[0]
        remaining_values = '\n'.join([f',{v}' for v in values_list[1:]])
        output = f"{comment_block}\nINSERT INTO `{table_name}` {column_def} VALUES {first_value}\n{remaining_values};\n"
    
    return output, table_name, record_count


def process_pasted_content(raw_text: str) -> list[tuple[str, str | None, int]]:
    """
    Process pasted INSERT statements (may contain multiple blocks separated by empty lines).
    
    Returns: List of (processed_content, table_name, record_count) tuples
    """
    results = []
    
    # Try to split by comment blocks (/* ... */)
    blocks = re.split(r'\n\s*\n(?=/\*)', raw_text)
    
    for block in blocks:
        block = block.strip()
        if block:
            processed, table_name, record_count = process_single_content(block)
            results.append((processed, table_name, record_count))
    
    return results


def process_folder(input_folder: str, output_folder: str, date_prefix: str, 
                   generate_combined: bool = False) -> dict:
    """
    Process all SQL files from input folder and write to output folder.
    
    Returns: Dictionary with results including processed files and any errors
    """
    results = {
        'processed': [],
        'errors': [],
        'combined_file': None
    }
    
    input_path = Path(input_folder)
    output_path = Path(output_folder)
    
    # Ensure output directory exists
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Get all .sql files
    sql_files = sorted(input_path.glob('*.sql'))
    
    if not sql_files:
        results['errors'].append(f"No .sql files found in {input_folder}")
        return results
    
    processed_contents = []
    
    for sql_file in sql_files:
        try:
            with open(sql_file, 'r', encoding='utf-8') as f:
                raw_content = f.read()
            
            processed_content, table_name, record_count = process_single_content(raw_content)
            
            # Write to output directory
            output_file = output_path / sql_file.name
            with open(output_file, 'w', encoding='utf-8', newline='\n') as f:
                f.write(processed_content)
            
            results['processed'].append({
                'input': sql_file.name,
                'output': output_file.name,
                'table': table_name,
                'records': record_count
            })
            
            if generate_combined and table_name:
                processed_contents.append((table_name, processed_content))
                
        except Exception as e:
            results['errors'].append(f"{sql_file.name}: {str(e)}")
    
    # Generate combined file if requested
    if generate_combined and processed_contents:
        combined_content = []
        for table_name, content in processed_contents:
            truncate_stmt = f"TRUNCATE TABLE {table_name};\r\n"
            combined_content.append(truncate_stmt)
            combined_content.append(content)
            combined_content.append("\r\n")
        
        combined_filename = f"{date_prefix} Combined.sql"
        combined_path = output_path / combined_filename
        
        with open(combined_path, 'w', encoding='utf-8', newline='') as f:
            f.write(''.join(combined_content))
        
        results['combined_file'] = combined_filename
    
    return results


def process_pasted_to_files(raw_text: str, output_folder: str, date_prefix: str,
                            generate_combined: bool = False) -> dict:
    """
    Process pasted INSERT statements and save to files.
    
    Returns: Dictionary with results including processed files and any errors
    """
    results = {
        'processed': [],
        'errors': [],
        'combined_file': None
    }
    
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)
    
    processed_items = process_pasted_content(raw_text)
    
    processed_contents = []
    
    for i, (processed_content, table_name, record_count) in enumerate(processed_items, 1):
        try:
            if table_name:
                filename = f"{date_prefix}_{i:02d}_{table_name}.sql"
            else:
                filename = f"{date_prefix}_{i:02d}_unknown.sql"
            
            output_file = output_path / filename
            with open(output_file, 'w', encoding='utf-8', newline='\n') as f:
                f.write(processed_content)
            
            results['processed'].append({
                'output': filename,
                'table': table_name,
                'records': record_count
            })
            
            if generate_combined and table_name:
                processed_contents.append((table_name, processed_content))
                
        except Exception as e:
            results['errors'].append(f"Block {i}: {str(e)}")
    
    # Generate combined file if requested
    if generate_combined and processed_contents:
        combined_content = []
        for table_name, content in processed_contents:
            truncate_stmt = f"TRUNCATE TABLE {table_name};\r\n"
            combined_content.append(truncate_stmt)
            combined_content.append(content)
            combined_content.append("\r\n")
        
        combined_filename = f"{date_prefix} Combined.sql"
        combined_path = output_path / combined_filename
        
        with open(combined_path, 'w', encoding='utf-8', newline='') as f:
            f.write(''.join(combined_content))
        
        results['combined_file'] = combined_filename
    
    return results


def process_files(file_paths: list, output_folder: str, date_prefix: str,
                  generate_combined: bool = False) -> dict:
    """
    Process a list of specific SQL files and write to output folder.
    
    Returns: Dictionary with results including processed files and any errors
    """
    results = {
        'processed': [],
        'errors': [],
        'combined_file': None
    }
    
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)
    
    if not file_paths:
        results['errors'].append("No files provided")
        return results
    
    processed_contents = []
    
    for file_path in file_paths:
        sql_file = Path(file_path)
        try:
            with open(sql_file, 'r', encoding='utf-8') as f:
                raw_content = f.read()
            
            processed_content, table_name, record_count = process_single_content(raw_content)
            
            # Write to output directory
            output_file = output_path / sql_file.name
            with open(output_file, 'w', encoding='utf-8', newline='\n') as f:
                f.write(processed_content)
            
            results['processed'].append({
                'input': sql_file.name,
                'output': output_file.name,
                'table': table_name,
                'records': record_count
            })
            
            if generate_combined and table_name:
                processed_contents.append((table_name, processed_content))
                
        except Exception as e:
            results['errors'].append(f"{sql_file.name}: {str(e)}")
    
    # Generate combined file if requested
    if generate_combined and processed_contents:
        combined_content = []
        for table_name, content in processed_contents:
            truncate_stmt = f"TRUNCATE TABLE {table_name};\r\n"
            combined_content.append(truncate_stmt)
            combined_content.append(content)
            combined_content.append("\r\n")
        
        combined_filename = f"{date_prefix} Combined.sql"
        combined_path = output_path / combined_filename
        
        with open(combined_path, 'w', encoding='utf-8', newline='') as f:
            f.write(''.join(combined_content))
        
        results['combined_file'] = combined_filename
    
    return results


def generate_excel_summary(processed_items: list, output_folder: str, date_prefix: str) -> str | None:
    """
    Generate an Excel summary file with table names and record counts.
    
    Args:
        processed_items: List of dicts with 'table' and 'records' keys
        output_folder: Where to save the Excel file
        date_prefix: Prefix for the filename
    
    Returns: Excel filename if successful, None otherwise
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return None  # openpyxl not installed
    
    if not processed_items:
        return None
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Count Summary"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="623a96", end_color="623a96", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="left", vertical="center")
    count_alignment = Alignment(horizontal="right", vertical="center")
    
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="FFF0DB", end_color="FFF0DB", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    ws['A1'] = 'table_name'
    ws['B1'] = 'count'
    
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws['A1'].border = thin_border
    
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill
    ws['B1'].alignment = header_alignment
    ws['B1'].border = thin_border
    
    # Data rows
    total_count = 0
    row = 2
    for item in processed_items:
        table_name = item.get('table', 'unknown')
        records = item.get('records', 0)
        total_count += records
        
        ws.cell(row=row, column=1, value=table_name)
        ws.cell(row=row, column=2, value=records)
        
        ws.cell(row=row, column=1).font = data_font
        ws.cell(row=row, column=1).alignment = data_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).font = data_font
        ws.cell(row=row, column=2).alignment = count_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        row += 1
    
    # Total row
    ws.cell(row=row, column=1, value='TOTAL')
    ws.cell(row=row, column=2, value=total_count)
    
    ws.cell(row=row, column=1).font = total_font
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).alignment = data_alignment
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2).font = total_font
    ws.cell(row=row, column=2).fill = total_fill
    ws.cell(row=row, column=2).alignment = count_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    
    # Save
    filename = f"{date_prefix} Count Summary.xlsx"
    filepath = Path(output_folder) / filename
    wb.save(filepath)
    
    return filename
